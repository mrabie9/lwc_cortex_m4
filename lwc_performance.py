import serial
import struct
import subprocess
import os
import time
import signal
import openpyxl as xl
from enum import Enum

# Get current directory
wdir = os.getcwd()

# Options
rebuild = False
xlwrite = True
# start_row = 14

# Board details 
board = "m4"

if board == "m4":
	xl_output = "Data/Data.xlsx"
	serial_port =  "COM6" 
elif board == "m7":
	xl_output = "Data/Data_m7.xlsx"
	serial_port =  "COM3" 	

# App details
algorithm = ''
# algorithms = ["ascon128", "ascon128Armv7", "Ascon128a", "ascon128aArmv7",
# 			  "Isapa128v20", "isapa128v20Armv7", "Isapa128av20", "isapa128av20Armv7",
# 			  "schwaemm256128v2", "schwaemm256128v2Armv7", "schwaemm256256v2", "schwaemm256256v2Armv7",
# 			  "TinyJambu", "tinyjambuOpt", "giftcofb128v1", "xoodyak", "romulusn", "romulusnOpt",
# 			  "elephant160v2", "grain128aeadv2", "photonbeetleaead128rate128v1"]
algorithms = ["ascon128"]
opts = ["_O3", "_O2", "_Os", "_O0"]
rows = [25, 14, 37, 3]
data_size = "12kB"

class AUT_col(Enum):
    ascon128 = 3
    ascon128Armv7 = 4
    ascon128a = 5
    ascon128aArmv7 = 6
    Isapa128v20 = 7
    isapa128v20Armv7 = 8
    Isapa128av20 = 9
    isapa128av20Armv7 = 10
    schwaemm256128v2 = 11
    schwaemm256128v2Armv7 = 12
    schwaemm256256v2 = 13
    schwaemm256256v2Armv7 = 14
    TinyJambu = 15
    tinyjambuOpt = 16
    giftcofb128v1 = 17
    xoodyak = 18
    romulusn = 19
    romulusnOpt = 20
    elephant160v2 = 21
    grain128aeadv2 = 22
    photonbeetleaead128rate128v1 = 23

start_col = AUT_col.ascon128.value

countdown_to_reset = 50

# Application runtime timeout
apprun = 120

def start_board(algorithm, opt):
	global countdown_to_reset
	prog_log = subprocess.run(wdir + "\openocd_" + board + ".sh " + algorithm + opt, shell=True, capture_output=True, text=True) # TODO: review
	if "Error" in prog_log.stdout or "Error" in prog_log.stderr or "Failed" in prog_log.stdout or "Failed" in prog_log.stderr:
		print(prog_log.stdout, '\n', prog_log.stderr)
		os._exit(10)
	print(prog_log.stdout, '\n', prog_log.stderr)
	clearBuffer()
	countdown_to_reset -= 1
	time.sleep(.1)

def signal_handler(signum, frame):
	print("Catched Ctrl+c")
	nucleo.close()
	global p
	try:
		os.waitpid(p.pid, os.WNOHANG)
	except:
		print("Nevermind...")
	print("Closed serial port")
	print(quit)
	os._exit(9)

def writexl(values, col, row):
    # open workbook
    workbook = xl.load_workbook(xl_output)
    ws = workbook['DWT - ' + board.upper()]

    # Write the variable names to the first col
    # for row, var in enumerate(variables, start=row):
    #     ws.cell(row=row, column=col-1, value=var)
    # row = row_start
    for row, var in enumerate(values, start=row):
        ws.cell(row=row, column=col, value=var)
    workbook.save(xl_output)

signal.signal(signal.SIGINT, signal_handler)

def float_to_hex(f):
    return struct.pack('<f', f)

def int_to_byte(f):
    return struct.pack('<B', f)

def uint64_to_hex(f):
    return struct.pack('<Q', f)

def int_to_hex(f):
    return struct.pack('<i', f)

def byte_to_int(f):
	if len(f) == 1:
		return struct.unpack('<B', f)[0]
	return -1 

def hex_to_ushort(f):
	if len(f) == 2:
		return struct.unpack('<H', f)[0]
	return -1 

def hex_to_float(f):
	# check if data has 4 bytes, needed by unpack
	if len(f) == 4:
		return struct.unpack("<f" ,f)[0]
	return -1 

def hex_to_double(f):
	if len(f) == 8:
		return struct.unpack("<d" ,f)[0]
	return -1 

def hex_to_uint64(f):
	if len(f) == 8:
		return struct.unpack("<Q" ,f)[0]
	return -1 

def hex_to_uint32(f):
	if len(f) == 4:
		return struct.unpack("<I" ,f)[0]
	return -1 

def hex_to_int(f):
	if len(f) == 4:
		return struct.unpack("<i" ,f)[0]
	return -1 

def hex_xor(f1, f2):
    return bytearray([a^b for a,b in zip(f1, f2)])

def clearBuffer():
	max_n_attempts = 50
	n_attempts = 1

	while (n_attempts <= max_n_attempts):
		a = nucleo.read(4)
		print("Attempting to clear buffer:")
		print(a)
		if (a == b''):
			print("Serial buffer is clean")
			break
		n_attempts += 1

	max_n_attempts += 1

	if n_attempts == max_n_attempts:
		print("Serial buffer is still not clean after attempt: ", n_attempts)
		print(quit)
		os._exit(17)


def sync():
	# print("Sending zero")
	nucleo.write(float_to_hex(0.0)) # Send 0
	val = nucleo.read(4)			# Receive 0 if successful

	# Check received value
	if(val == b''):
		print("Error: Sent 0 but received null")
		return 1
	elif(hex_to_float(val) != 0.0):
		print("Error: Sent 0 but received ", hex_to_float(val))
		return 1
	else:
		return 0
	
try:
	nucleo = serial.Serial(serial_port, 115200, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout = 5)
except:
	print("Problem opening serial port")
	os._exit(16)



rebuild_output_filename = wdir + r"\rebuild_output_" + data_size
if rebuild:
	print("Rebuilding All. Please wait...")
	rebuild_log = subprocess.run(wdir + r"\rebuild_all.sh " + rebuild_output_filename, shell=True, capture_output=True, text=True) # TODO: review
	if "Error" in rebuild_log.stdout or "Error" in rebuild_log.stderr or "Failed" in rebuild_log.stdout or "Failed" in rebuild_log.stderr:
		print(rebuild_log.stdout, '\n', rebuild_log.stderr)
		os._exit(10)

# filename = wdir + "\Opt_" + board + "_" + data_size + ".txt"
# print(filename)
# f = open(filename, "w") # clear file first
# f.close()
# f = open(filename, "a")
# if f.closed:
# 	print("File not open!")

col = start_col

for opt in opts:
	start_row = rows[opts.index(opt)]
	for algorithm in algorithms:
		if 'O0' in opt and ('Armv7' or 'Opt') in algorithm:
			continue
		
		start_board(algorithm, opt) # flash board
		print("Starting experiment")
		print(algorithm + opt + ":")

		# Main loop
		try:
			if countdown_to_reset == 0:
				os._exit(10)

			# Sync with app
			if sync() == 1:
				continue
			AUT_start_time = time.time()

			## Wait for enc and dec to finish
			nucleo.timeout = apprun
			nucleo.read(4)
			nucleo.timeout = 2

			## Python timer
			runtime_py = round(time.time() - AUT_start_time, 6)
			print(runtime_py)
			nucleo.read(4)

			## Start collecting data from AUT
			if sync() == 1:
				continue
			runtime_e = hex_to_float(nucleo.read(4))
			if sync() == 1:
				continue
			runtime_d = hex_to_float(nucleo.read(4))
			
			# Get ouptput results
			if sync() == 1:
				continue
			output_e = hex_to_double(nucleo.read(8))
			if sync() == 1:
				continue
			output_d = hex_to_double(nucleo.read(8))

			# Get error counter
			if sync() == 1:
				continue
			sum = hex_to_uint32(nucleo.read(4))

			## Print results
			print("Runtime E: %f s" % runtime_e)
			print("Runtime D: %f s" % runtime_d)		
			print("Output E: ", output_e)
			print("Output D: ", output_d)
			print("Err Cnt:  ", sum)
			print("Runtime Py: ", runtime_py , " s")
			
			## Write results to text file
			# f.write(algorithm + ": \n")
			# f.write("\tRuntime E:\t\t%f s\n\t" % runtime_e)
			# f.write("Runtime D:\t\t%f s\n\t" % runtime_d)
			# f.write("Output E:\t\t%.1f\n\t" % output_e)
			# f.write("Output D:\t\t%.1f\n\t" % output_d)
			# f.write("Err Cnt:\t\t%.1f\n\t" % sum)
			# f.write("Runtime Py:\t\t" + str(runtime_py) + " s\n\n")

			## Write to spreadsheet
			if xlwrite: 
				# find col number
				for col in list(AUT_col):
					if algorithm == col.name:
						curr_col = col.value
				values = [runtime_e, runtime_d, output_e,  output_d, sum, runtime_py]
				writexl(values, curr_col, start_row)

			countdown_to_reset = 5
			time.sleep(.05)

		except Exception as e:
			print("Shouldnt be here->", e)
			start_board()
			continue

# f.close()