import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
import time
from natsort import natsorted
import openpyxl as xl
from enum import Enum

global x_start_e, x_stop_e, x_start_d, x_stop_d, n_loop, energy_e, energy_d
global i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated, t_calibrations
global n_lcutoff_points, l_cut_off, n_hcutoff_points, h_cut_off, calibrated
global values 

variables = ['n_loop', 'Average encryption time', 'Average decryption time', 'Average encryption energy', 'Average decryption energy',
             'Maximum encryption current', 'Average encryption current', 'Minimum encryption current', 
             'Maximum decryption current', 'Average decryption current', 'Minimum decryption current',
             'Calibrations', 'Calibration time(s)']

# values = [n_loop, t_avg_e, t_avg_d,  energy_e, energy_d,
#           i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated,', '.join(map(str, t_calibrations))]


# Contains column value for the AUT
class AUT(Enum):
    ascon128_enc_dec_1000x = 3
    ascon128Armv7_enc_dec_1000x = 4
    ascon128a_enc_dec_1000x = 5
    ascon128aArmv7_enc_dec_1000x = 6
    isapa128_enc_dec_500x = 7
    isapa128Armv7_enc_dec_500x = 8
    isapa128a_enc_dec_500x = 9
    isapa128aArmv7_enc_dec_500x = 10
    sparkle128_enc_dec_1000x = 11
    sparkle128Armv7_enc_dec_1000x = 12
    sparkle256_enc_dec_1000x = 13
    sparkle256Armv7_enc_dec_1000x = 14
    tinyjambu_enc_dec_1000x = 15
    tinyjambuOpt_enc_dec_1000x = 16
    giftc_enc_dec_200x = 17
    xoodyak_enc_dec_1000x = 18
    romulusn_enc_dec_50x = 19
    romulusnOpt_enc_dec_500x = 20
    eleph_enc_dec_10x = 21
    grain_enc_dec_10x = 22
    photon_enc_dec_15x = 23

# Start col for xlsx write
col = AUT.ascon128_enc_dec_1000x.value 

# Start/end times dict
# data_obtained = False#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# energy_calc = False#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/O0/ascon128a2_enc_dec_1000x/"
# output = "../Data/Power/python_plots/00_Output/02_run2/O0/ascon128a2_enc_dec_1000x.txt"
# timestamps = {
#     ## O0 
#     # 'ascon128_enc_dec_1000x': {'start_e': 1.096410, 'stop_e':183.097190 , 'start_d': 186.102970, 'stop_d':368.140520},
#     # 'ascon128a_enc_dec_1000x': {'start_e': 0.868760, 'stop_e':126.672980, 'start_d': 129.678520 , 'stop_d':255.602330},
#     # 'giftc_enc_dec_200x': {'start_e': 0.979360, 'stop_e': 234.564920, 'start_d': 237.570210, 'stop_d': 471.166760},
#     # 'isapa128_enc_dec_500x': {'start_e': 1.106770, 'stop_e': 336.044850 , 'start_d': 339.051670, 'stop_d': 674.354140},
#     # 'isapa128a_enc_dec_500x': {'start_e': 1.095630, 'stop_e': 263.663310, 'start_d': 266.669700 , 'stop_d': 529.252050},
#     # 'sparkle128_enc_dec_1000x': {'start_e': 1.064100, 'stop_e': 84.239800, 'start_d': 87.245910, 'stop_d': 171.738190},
#     # 'sparkle256_enc_dec_1000x': {'start_e': 1.142160, 'stop_e': 116.325410, 'start_d': 119.331090, 'stop_d': 235.864220 },
#     # 'tinyjambu_enc_dec_1000x': {'start_e': 1.019490, 'stop_e': 158.006720, 'start_d': 161.009910, 'stop_d': 318.775550},
#     # 'xoodyak_enc_dec_1000x': {'start_e': 1.34246, 'stop_e': 282.81752, 'start_d': 285.82417, 'stop_d':  567.05898},
#     # 'eleph_enc_dec_10x': {'start_e': 5.469650, 'stop_e': 227.465460, 'start_d': 230.472160, 'stop_d': 452.353150 },
#     # 'grain_enc_dec_10x': {'start_e': 1.445850, 'stop_e': 183.564560, 'start_d': 186.567760, 'stop_d': 366.392920},
#     # 'photon_enc_dec_15x': {'start_e': 1.13041, 'stop_e': 203.55816, 'start_d': 206.56335, 'stop_d':409.04147},
#     # 'romulusn_enc_dec_50x': {'start_e': 1.180800, 'stop_e': 182.506460, 'start_d': 185.512820, 'stop_d': 367.872190},
# }
data_obtained = True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
energy_calc = True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/O0/isapa128a_enc_dec_500x/"
# output = "../Data/Power/python_plots/00_Output/02_run2/O0/isapa128a_enc_dec_500x.txt"
row = 3
timestamps = {
#     ## Run 2 start: O0 (with sync) 
#     # 'O0': {
#     #     'row' : 3,
#     #     'ascon128_enc_dec_1000x': {'start_e': 1.35464, 'stop_e': 183.32447, 'start_d': 186.32992, 'stop_d': 368.37830},
#     #     'ascon128a_enc_dec_1000x': {'start_e': 1.37838, 'stop_e': 127.13827, 'start_d': 130.14297, 'stop_d': 256.05088},
#     #     'isapa128_enc_dec_500x': {'start_e': 1.27187, 'stop_e': 336.41365, 'start_d': 339.41733, 'stop_d': 674.50535},
#     #     'isapa128a_enc_dec_500x': {'start_e': 1.29430, 'stop_e': 263.7350, 'start_d': 266.73994, 'stop_d': 529.2357},
#     #     'sparkle128_enc_dec_1000x': {'start_e': 1.19860, 'stop_e': 84.34733, 'start_d': 87.35280, 'stop_d': 171.83800},
#     #     'sparkle256_enc_dec_1000x':{'start_e': 1.30745, 'stop_e': 116.48676, 'start_d': 119.49210, 'stop_d': 236.04129},
#     #     'tinyjambu_enc_dec_1000x': {'start_e': 1.27229, 'stop_e': 159.12436, 'start_d': 162.12985, 'stop_d': 320.05205},
#     #     'giftc_enc_dec_200x':{'start_e': 1.28366, 'stop_e': 234.75802, 'start_d': 237.76210, 'stop_d': 471.14448},
#     #     'xoodyak_enc_dec_1000x': {'start_e': 1.22420, 'stop_e': 282.73972, 'start_d': 285.74527, 'stop_d': 566.93847},
#     #     'romulusn_enc_dec_50x': {'start_e': 1.16678, 'stop_e': 183.72276, 'start_d': 186.73369, 'stop_d': 369.45473}
#     #     'eleph_enc_dec_10x': {'start_e': 1.32715, 'stop_e': 223.14892, 'start_d': 226.15294, 'stop_d': 448.04365},
#     #     'grain_enc_dec_10x': {'start_e': 1.21045, 'stop_e': 183.78595, 'start_d': 186.7954, 'stop_d': 367.01580},
#     #     'photon_enc_dec_15x': {'start_e': 1.26331, 'stop_e': 203.95343, 'start_d': 206.96480, 'stop_d': 409.72025},
#     # },

#     'O2': {
#         'row' : 107,#20,
#         'ascon128_enc_dec_1000x': {'start_e': 1.32980, 'stop_e': 36.96500, 'start_d': 39.97337, 'stop_d': 75.42982},
#         'ascon128Armv7_enc_dec_1000x': {'start_e': 1.41271, 'stop_e': 27.90305, 'start_d': 30.90838, 'stop_d': 57.30829},
#         'ascon128a_enc_dec_1000x': {'start_e': 1.33708, 'stop_e': 27.43752, 'start_d': 30.44534, 'stop_d': 56.61588},
#         'ascon128aArmv7_enc_dec_1000x':{'start_e': 1.35539, 'stop_e': 18.93792, 'start_d': 21.94486, 'stop_d': 39.41609},
#         'isapa128_enc_dec_500x': {'start_e': 1.22425, 'stop_e': 72.29740, 'start_d': 75.30481, 'stop_d': 146.38173},
#         'isapa128Armv7_enc_dec_500x': {'start_e': 1.45076, 'stop_e': 46.52175, 'start_d': 49.52730, 'stop_d': 94.60091},
#         'isapa128a_enc_dec_500x': {'start_e': 1.28838, 'stop_e': 63.50118, 'start_d': 66.50847, 'stop_d': 128.72280},
#         'isapa128aArmv7_enc_dec_500x': {'start_e': 1.38490, 'stop_e': 33.95361, 'start_d': 36.95904, 'stop_d': 69.53307},
#         'sparkle128_enc_dec_1000x': {'start_e': 1.33705, 'stop_e': 16.94720, 'start_d': 19.95426, 'stop_d': 36.23320},
#         'sparkle128Armv7_enc_dec_1000x': {'start_e': 1.47810, 'stop_e': 12.55021, 'start_d': 15.55501, 'stop_d': 26.62793},
#         'sparkle256_enc_dec_1000x':{'start_e': 1.21061, 'stop_e': 22.15091, 'start_d': 25.15783, 'stop_d': 46.61738},
#         'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.37846, 'stop_e': 18.40788, 'start_d': 21.41232, 'stop_d': 38.41758},
#         'tinyjambu_enc_dec_1000x': {'start_e': 1.30960, 'stop_e': 40.70770, 'start_d': 43.70199, 'stop_d': 83.04224},
#         'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.40447, 'stop_e': 25.60081, 'start_d': 28.60497, 'stop_d': 52.72808},
#         'giftc_enc_dec_200x':{'start_e': 1.33015, 'stop_e': 83.28543, 'start_d': 86.29315, 'stop_d': 168.29000},
#         'xoodyak_enc_dec_1000x': {'start_e': 1.14605, 'stop_e': 75.25774, 'start_d': 78.26672, 'stop_d': 152.21332},
#         'romulusn_enc_dec_50x': {'start_e': 2.00606, 'stop_e': 30.47147, 'start_d': 33.48210, 'stop_d': 61.99735},
#         'romulusnOpt_enc_dec_50x': {'start_e': 1.33416, 'stop_e': 6.05559, 'start_d': 9.06258, 'stop_d': 13.77557},
#         'eleph_enc_dec_10x': {'start_e': 1.16114, 'stop_e': 53.6765, 'start_d': 56.68451, 'stop_d': 109.20734 },
#         'grain_enc_dec_10x': {'start_e': 1.22906, 'stop_e': 45.16622, 'start_d': 48.17469, 'stop_d': 91.87389},
#         'photon_enc_dec_15x': {'start_e': 1.17107, 'stop_e': 69.71969, 'start_d': 72.72957, 'stop_d': 141.30102},
#     },

#     'O3':{
#         ## O3 (with sync)
#         'row' : 108,# 37,
#         'ascon128_enc_dec_1000x': {'start_e': 1.61527, 'stop_e': 32.52472, 'start_d': 35.53011, 'stop_d': 66.53474},
#         'ascon128Armv7_enc_dec_1000x': {'start_e': 1.38153, 'stop_e': 27.83631, 'start_d': 30.84171, 'stop_d': 57.19952},
#         'ascon128a_enc_dec_1000x': {'start_e': 1.32484, 'stop_e': 22.90531, 'start_d': 25.91182, 'stop_d': 47.10141}, 
#         'ascon128aArmv7_enc_dec_1000x':{'start_e': 1.48811, 'stop_e': 19.07184, 'start_d': 22.07951, 'stop_d': 39.55043},
#         'isapa128_enc_dec_500x': {'start_e': 1.35027, 'stop_e': 50.71819, 'start_d': 53.72424, 'stop_d': 103.09404},
#         'isapa128Armv7_enc_dec_500x': {'start_e': 1.20896, 'stop_e': 46.11353, 'start_d': 49.11976, 'stop_d': 94.02679},
#         'isapa128a_enc_dec_500x': {'start_e': 1.40207, 'stop_e': 42.20895, 'start_d': 45.21585, 'stop_d': 86.02839},
#         'isapa128aArmv7_enc_dec_500x': {'start_e': 1.44101, 'stop_e': 33.81900, 'start_d': 36.82403, 'stop_d': 69.11772},
#         'sparkle128_enc_dec_1000x': {'start_e': 1.18393, 'stop_e': 17.27858, 'start_d': 20.28510, 'stop_d': 37.15058}, 
#         'sparkle128Armv7_enc_dec_1000x': {'start_e': 1.39177, 'stop_e': 12.58562, 'start_d': 15.59153, 'stop_d': 26.78508},
#         'sparkle256_enc_dec_1000x':{'start_e': 0.99120, 'stop_e': 33.05594, 'start_d': 36.06287, 'stop_d': 68.77249},
#         'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.41073, 'stop_e': 18.66018, 'start_d': 21.66667, 'stop_d': 38.93965},
#         'tinyjambu_enc_dec_1000x': {'start_e': 1.36392, 'stop_e': 33.29568, 'start_d': 36.30292, 'stop_d': 68.35199},
#         'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.45441, 'stop_e': 25.60803, 'start_d': 28.61335, 'stop_d': 52.76988},
#         'xoodyak_enc_dec_1000x': {'start_e': 1.44806, 'stop_e': 33.57857, 'start_d': 36.58487, 'stop_d': 68.34107},
#         'giftc_enc_dec_200x': {'start_e': 1.32349, 'stop_e': 93.03627, 'start_d': 96.04343, 'stop_d': 187.75343},
#         'romulusn_enc_dec_50x': {'start_e': 1.25101, 'stop_e': 20.42448, 'start_d': 23.43189, 'stop_d': 42.62002},
#         'romulusnOpt_enc_dec_50x': {'start_e': 1.44879, 'stop_e': 6.17670, 'start_d': 9.18398, 'stop_d': 13.89927},
#         'eleph_enc_dec_10x': {'start_e': 1.34278, 'stop_e': 36.84270, 'start_d': 39.84887, 'stop_d': 75.34979},
#         'grain_enc_dec_10x': {'start_e': 1.31027, 'stop_e': 43.01828, 'start_d': 46.02507, 'stop_d': 87.55085}, 
#         'photon_enc_dec_15x': {'start_e': 1.26719, 'stop_e': 39.84326, 'start_d': 42.85026, 'stop_d': 81.38094},
#     },

    'Os':{
        ## Os (with sync)
        'row' : 54,
        'ascon128_enc_dec_1000x': {'start_e': 1.38023, 'stop_e': 29.11083, 'start_d': 31.93298, 'stop_d': 59.51844}, 
        'ascon128Armv7_enc_dec_1000x': {'start_e': 1.28564, 'stop_e': 28.11030, 'start_d': 31.11579, 'stop_d': 57.81297}, 
        'ascon128a_enc_dec_1000x': {'start_e': 1.30820, 'stop_e': 21.86800, 'start_d': 24.87548, 'stop_d': 45.48583},
        'ascon128aArmv7_enc_dec_1000x': {'start_e': 1.46410, 'stop_e': 19.09462, 'start_d': 22.10305, 'stop_d': 39.71577},
        'isapa128_enc_dec_500x': {'start_e': 0.66834, 'stop_e': 86.14700, 'start_d': 89.15498, 'stop_d': 174.64830},
        'isapa128Armv7_enc_dec_500x': {'start_e': 1.47178, 'stop_e': 47.01213, 'start_d': 50.01704, 'stop_d': 95.59974},
        'isapa128a_enc_dec_500x': {'start_e': 1.28150, 'stop_e': 74.92620, 'start_d': 77.93461, 'stop_d': 151.55463},
        'isapa128aArmv7_enc_dec_500x': {'start_e': 1.38657, 'stop_e': 34.47951, 'start_d': 37.48394, 'stop_d': 70.58174},
        'sparkle128_enc_dec_1000x': {'start_e': 1.40150, 'stop_e': 18.32346, 'start_d': 21.33160, 'stop_d': 39.01026},
        'sparkle128Armv7_enc_dec_1000x': {'start_e': 1.31934, 'stop_e': 12.60383, 'start_d': 15.60876, 'stop_d': 26.92872},
        'sparkle256_enc_dec_1000x': {'start_e': 1.43601, 'stop_e': 24.51053, 'start_d': 27.51840, 'stop_d': 51.51285}, 
        'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.23802, 'stop_e': 18.41306, 'start_d': 21.41884, 'stop_d': 38.55277},
        'tinyjambu_enc_dec_1000x': {'start_e': 1.41110, 'stop_e':  46.64048, 'start_d': 49.65009, 'stop_d': 94.85015},
        'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.47501, 'stop_e': 27.72689, 'start_d': 30.73108, 'stop_d': 56.94882},
        'giftc_enc_dec_200x':{'start_e': 1.44594, 'stop_e': 82.20315, 'start_d': 85.21073, 'stop_d': 165.98799},
        'xoodyak_enc_dec_1000x': {'start_e': 1.38990, 'stop_e': 84.06259, 'start_d': 87.07055, 'stop_d': 169.67377}, 
        'romulusn_enc_dec_50x': {'start_e': 1.40887, 'stop_e': 36.00575, 'start_d': 39.01242, 'stop_d': 73.67868},
        'romulusnOpt_enc_dec_50x': {'start_e': 1.38247, 'stop_e': 6.10803, 'start_d': 9.11501, 'stop_d': 13.84141},
        'eleph_enc_dec_10x': {'start_e': 1.32568, 'stop_e': 73.83440, 'start_d': 76.84218, 'stop_d': 149.35690},
        'grain_enc_dec_10x': {'start_e': 1.27562, 'stop_e': 49.83802, 'start_d': 52.84632, 'stop_d': 101.17227},
        'photon_enc_dec_15x': {'start_e': 1.00620, 'stop_e': 75.79236, 'start_d': 78.79877, 'stop_d': 153.61510},
    },  
    'Os2':{
        ## Os (with sync)
        'row' : 74,
        'ascon128_enc_dec_1000x': {'start_e': 1.29531, 'stop_e': 28.83720, 'start_d': 31.84077, 'stop_d': 59.39126}, 
        'ascon128Armv7_enc_dec_1000x': {'start_e': 1.61856, 'stop_e': 28.43439, 'start_d': 31.43902, 'stop_d': 58.12702}, 
        'ascon128a_enc_dec_1000x': {'start_e': 1.55824, 'stop_e': 22.10180, 'start_d': 25.10603, 'stop_d': 45.69724}, 
        'ascon128aArmv7_enc_dec_1000x': {'start_e': 1.12357, 'stop_e': 18.73460, 'start_d': 21.73778, 'stop_d': 39.33007}, 
        'isapa128_enc_dec_500x': {'start_e': 1.39975, 'stop_e': 85.70521, 'start_d': 88.70845, 'stop_d': 173.11105}, 
        'isapa128a_enc_dec_500x': {'start_e': 1.48134, 'stop_e': 74.33800, 'start_d': 77.33984, 'stop_d': 150.19907}, 
        'sparkle128_enc_dec_1000x': {'start_e': 1.52926, 'stop_e': 18.44535, 'start_d': 21.44764, 'stop_d': 39.09003}, 
        'sparkle128Armv7_enc_dec_1000x': {'start_e': 1.57333, 'stop_e': 12.85415, 'start_d': 15.85581, 'stop_d': 27.10567}, 
        'sparkle256_enc_dec_1000x': {'start_e': 1.63753, 'stop_e': 24.66953, 'start_d': 27.67147, 'stop_d': 51.62081}, 
        'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.49618, 'stop_e': 18.65213, 'start_d': 21.65428, 'stop_d': 38.76583}, 
        'tinyjambu_enc_dec_1000x': {'start_e': 1.56627, 'stop_e': 46.70575, 'start_d': 49.70775, 'stop_d': 94.80927}, 
        'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.51190, 'stop_e': 27.73796, 'start_d': 30.73904, 'stop_d': 56.93416}, 
        'giftc_enc_dec_200x': {'start_e': 1.51069, 'stop_e': 82.08676, 'start_d': 85.08783, 'stop_d': 165.69001}, 
        'xoodyak_enc_dec_1000x': {'start_e': 1.47387, 'stop_e': 83.96918, 'start_d': 86.97155, 'stop_d': 169.38805}, 
        'romulusn_enc_dec_50x': {'start_e': 1.16773, 'stop_e': 35.72356, 'start_d': 38.72515, 'stop_d': 73.35093}, 
        'romulusnOpt_enc_dec_500x': {'start_e': 1.56701, 'stop_e': 6.28361, 'start_d': 9.28512, 'stop_d': 14.00304}, 
        'eleph_enc_dec_10x': {'start_e': 1.55702, 'stop_e': 73.89897, 'start_d': 76.90055, 'stop_d': 149.23618}, 
        'grain_enc_dec_10x': {'start_e': 1.38815, 'stop_e': 49.83820, 'start_d': 52.83938, 'stop_d': 101.07402}, 
        'photon_enc_dec_15x': {'start_e': 1.44612, 'stop_e': 76.07520, 'start_d': 79.07596, 'stop_d': 153.70561},  
    }
}
# timestamps = {
#     'Os':{
#         ## Os (with sync)
#         'row' : 74,
#         'ascon128_enc_dec_1000x': {'start_e': 1.29531, 'stop_e': 28.83720, 'start_d': 31.84077, 'stop_d': 59.39126}, 
#         'ascon128Armv7_enc_dec_1000x': {'start_e': 1.61856, 'stop_e': 28.43439, 'start_d': 31.43902, 'stop_d': 58.12702}, 
#         'ascon128a_enc_dec_1000x': {'start_e': 1.55824, 'stop_e': 22.10180, 'start_d': 25.10603, 'stop_d': 45.69724}, 
#         'ascon128aArmv7_enc_dec_1000x': {'start_e': 1.12357, 'stop_e': 18.73460, 'start_d': 21.73778, 'stop_d': 39.33007}, 
#         'isapa128_enc_dec_500x': {'start_e': 1.39975, 'stop_e': 85.70521, 'start_d': 88.70845, 'stop_d': 173.11105}, 
#         'isapa128a_enc_dec_500x': {'start_e': 1.48134, 'stop_e': 74.33800, 'start_d': 77.33984, 'stop_d': 150.19907}, 
#         'sparkle128_enc_dec_1000x': {'start_e': 1.52926, 'stop_e': 18.44535, 'start_d': 21.44764, 'stop_d': 39.09003}, 
#         'sparkle128Armv7_enc_dec_1000x': {'start_e': 1.57333, 'stop_e': 12.85415, 'start_d': 15.85581, 'stop_d': 27.10567}, 
#         'sparkle256_enc_dec_1000x': {'start_e': 1.63753, 'stop_e': 24.66953, 'start_d': 27.67147, 'stop_d': 51.62081}, 
#         'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.49618, 'stop_e': 18.65213, 'start_d': 21.65428, 'stop_d': 38.76583}, 
#         'tinyjambu_enc_dec_1000x': {'start_e': 1.56627, 'stop_e': 46.70575, 'start_d': 49.70775, 'stop_d': 94.80927}, 
#         'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.51190, 'stop_e': 27.73796, 'start_d': 30.73904, 'stop_d': 56.93416}, 
#         'giftc_enc_dec_200x': {'start_e': 1.51069, 'stop_e': 82.08676, 'start_d': 85.08783, 'stop_d': 165.69001}, 
#         'xoodyak_enc_dec_1000x': {'start_e': 1.47387, 'stop_e': 83.96918, 'start_d': 86.97155, 'stop_d': 169.38805}, 
#         'romulusn_enc_dec_50x': {'start_e': 1.16773, 'stop_e': 35.72356, 'start_d': 38.72515, 'stop_d': 73.35093}, 
#         'romulusnOpt_enc_dec_500x': {'start_e': 1.56701, 'stop_e': 6.28361, 'start_d': 9.28512, 'stop_d': 14.00304}, 
#         'eleph_enc_dec_10x': {'start_e': 1.55702, 'stop_e': 73.89897, 'start_d': 76.90055, 'stop_d': 149.23618}, 
#         'grain_enc_dec_10x': {'start_e': 1.38815, 'stop_e': 49.83820, 'start_d': 52.83938, 'stop_d': 101.07402}, 
#         'photon_enc_dec_15x': {'start_e': 1.44612, 'stop_e': 76.07520, 'start_d': 79.07596, 'stop_d': 153.70561}, 
#     } 
# }
# data_obtained = False#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# energy_calc = False#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/O0/ascon128a2_enc_dec_1000x/"
# output = "../Data/Power/python_plots/00_Output/02_run2/O0/ascon128a2_enc_dec_1000x.txt"
# timestamps = {
#     ### 'Random' second sets
#     # 'photon2_enc_dec_15x': {'start_e': 0.78681, 'stop_e': 202.93627, 'start_d': 205.93713, 'stop_d': 407.90221},
#     # 'ascon128a_ngnd_enc_dec_1000x': {'start_e': 0., 'stop_e': , 'start_d': , 'stop_d': }
# }

plot_all_bool = False
opt = "Os"
run3 = "03_Run3"
run2 = "02_Run2"

run=run2

# apps = ["ascon128_enc_dec_1000x", "ascon128Armv7_enc_dec_1000x", "ascon128a_enc_dec_1000x", "ascon128aArmv7_enc_dec_1000x"]
# apps = ["isapa128_enc_dec_500x", "isapa128Armv7_enc_dec_500x", "isapa128a_enc_dec_500x", "isapa128aArmv7_enc_dec_500x", ]
# apps = [ "sparkle128_enc_dec_1000x", "sparkle128Armv7_enc_dec_1000x", "sparkle256_enc_dec_1000x", "sparkle256Armv7_enc_dec_1000x"]
# apps = ["tinyjambu_enc_dec_1000x", "tinyjambuOpt_enc_dec_1000x", "giftc_enc_dec_200x", "xoodyak_enc_dec_1000x",]
# apps = [ "romulusn_enc_dec_50x","romulusnOpt_enc_dec_500x","eleph_enc_dec_10x"]
apps = [ "grain_enc_dec_10x", "photon_enc_dec_15x"]

# apps  = ["ascon128aArmv7_enc_dec_1000x", "romulusnOpt_enc_dec_50x"]
def plot_all():
    dfs = []
    t = time.time()
    for app in apps:
        print("Reading ", app, apps.index(app))
        if "isapa128" in app and "Armv7" in app and "Os" in opt:
            continue
        data_dir = "../Data/Power/python_plots/02_csv_dir/"+ run + "/"+ opt + "/" + app + "/"
        files = os.listdir(data_dir)

        # Filter out only CSV files
        csv_files = [file for file in files if (file.endswith('.csv') and not "summary" in file)]
        csv_files = natsorted(csv_files)

        
        for file_name in csv_files:
            file_path = os.path.join(data_dir, file_name)
            # Read the CSV file into a DataFrame
            df = pd.read_csv(file_path, delimiter=';', names=[app + "x", app + "y"])
            dfs.append(df)

        # Concatenate all DataFrames into a single DataFrame ignore index HAS to be true
        df = pd.concat(dfs, ignore_index=True)
        # x_values[apps.index(app)] = df['x']/1e3
        # y_values[apps.index(app)] = df['y']/1e6
    print(time.time() - t)
    for app in apps:
        if "isapa128" in app and "Armv7" in app and "Os" in opt:
            continue
        t = time.time()
        x_values= df[app + 'x']/1e3
        y_values = df[app + 'y']/1e6
        # Plotting the graph
        plt.plot(x_values, y_values, linestyle='-') 
        # plt.figure()
        plt.xlabel('Time (s)') 
        plt.ylabel('Current (A)') 
        plt.title(app) 
        # plt.text(4.28073, 0.047, "Start of program", fontsize=12, ha='center', va='center', color='black')
        plt.grid(True)
        print(time.time() - t)
        plt.show()

def main(ts_dict):
    global x_start_e, x_stop_e, x_start_d, x_stop_d, n_loop, energy_e, energy_d
    global i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated, t_calibrations
    global n_lcutoff_points, l_cut_off, n_hcutoff_points, h_cut_off, calibrated
    global values
    # List all files in the directory
    files = os.listdir(data_dir)

    # Filter out only CSV files
    csv_files = [file for file in files if (file.endswith('.csv') and not "summary" in file)]
    csv_files = natsorted(csv_files)
    # print(csv_files)

    # Loop through each CSV file and read its contents
    dfs = []
    for file_name in csv_files:
        file_path = os.path.join(data_dir, file_name)
        # Read the CSV file into a DataFrame
        df = pd.read_csv(file_path, delimiter=';', names=["x", "y"])
        dfs.append(df)

    # Concatenate all DataFrames into a single DataFrame ignore index HAS to be true
    df = pd.concat(dfs, ignore_index=True)
    # Assuming your CSV file has columns named 'x' and 'y', change them accordingly if they are different
    x_values = df['x']/1e3
    y_values = df['y']/1e6
    # print(x_values)

    # Check for calibrations
    n_calibrated = len(y_values[y_values>1])
    t_calibrations = y_values[y_values>1].index
    calibrated = bool(n_calibrated)
    # print("Number of calibrations: ", n_calibrated)

    # remove calibration current samples
    l_cut_off = []
    h_cut_off = 1
    n_hcutoff_points =  len(y_values[y_values>h_cut_off])
    # print("Calibrated: ", calibrated)
    if len(y_values[y_values<0.005]) < 10:
        l_cut_off = 0.005
        n_lcutoff_points =  len(y_values[y_values<l_cut_off])
        x_values = x_values[y_values<1]
        x_values = x_values[y_values>0.005]
        y_values = y_values[y_values<1]
        y_values = y_values[y_values>0.005]
        # print("Removed " + str(n_lcutoff_points) + " at 0.005 A cut-off")
        # print("Removed " + str(n_hcutoff_points) + " above " + str(n_hcutoff_points) + "A")
    else:
        l_cut_off = 0.001
        n_lcutoff_points =  len(y_values[y_values<l_cut_off])
        x_values = x_values[y_values<1]
        x_values = x_values[y_values>0.001]
        y_values = y_values[y_values<1]
        y_values = y_values[y_values>0.001]
        # print("Removed " + str(n_lcutoff_points) + " below " + str(l_cut_off) + "A")
        # print("Removed " + str(n_hcutoff_points) + " above " + str(h_cut_off) + "A")

    ## get start/end time - for report
    # specific_x_val = 3.99519
    # index_of_x = x_values[x_values== specific_x_val].index[0]
    # x_values = x_values[index_of_x:]
    # y_values = y_values[index_of_x:]
    # annotate_x = x_values[index_of_x]
    # annotate_y = round(y_values[index_of_x], 4)
    # # Annotate the point with its coordinates
    # plt.annotate(f'({annotate_x}, {annotate_y})', xy=(annotate_x, annotate_y), xytext=(annotate_x-2e-5, annotate_y+1e-4))
    # plt.xlim(x_values[index_of_x]-9e-5,x_values[index_of_x]+5e-5 )
    # plt.gca().yaxis.set_ticks_position('right')
    # plt.gca().yaxis.set_label_position('right')

    # Figure out n_loop
    if "1000" in data_dir:
        n_loop = 1000
    elif "500" in data_dir:
        n_loop = 500
    elif "200" in data_dir:
        n_loop = 200
    elif "100" in data_dir:
        n_loop = 100
    elif "50" in data_dir:
        n_loop = 50
    elif "15" in data_dir:
        n_loop = 15
    elif "10" in data_dir:
        n_loop = 10
    else: 
        print("N_loop not recognised")
        os._exit

    if energy_calc:
        ## try nested dict
        x_start_e = 0
        x_stop_e = 0
        x_start_d = 0
        x_stop_d = 0

        for app, data in ts_dict.items():
            if app in data_dir:
                x_start_e = round(data['start_e'],5)
                x_stop_e = round(data['stop_e'],5)
                x_start_d = round(data['start_d'],5)
                x_stop_d = round(data['stop_d'],5)

        # energy calculation - encryption
        # x_start_e = 1.180800
        # x_stop_e = 182.506460
        id_x_start_e = x_values[round(x_values,5)== x_start_e].index[0]
        # print("Start enc time: ", x_values[id_x_start_e])
        id_x_stop_e = x_values[round(x_values,5)== x_stop_e].index[0]
        i_e_avg = round(y_values[id_x_start_e:id_x_stop_e].mean(),5)
        i_e_max = round(y_values[id_x_start_e:id_x_stop_e].max(),5)
        i_e_min = round(y_values[id_x_start_e:id_x_stop_e].min(),5)
        energy_e = i_e_avg * 3.3 * (x_stop_e-x_start_e)/n_loop
        energy_e = round(energy_e,5)
        # print("Average energy - encryption: ", energy_e, (x_stop_e-x_start_e)/n_loop)

        # energy calculation - decryption
        # x_start_d = 185.512820
        # x_stop_d = 367.872190
        id_x_start_d = x_values[round(x_values,5)== x_start_d].index[0]
        id_x_stop_d = x_values[round(x_values,5)== x_stop_d].index[0]
        # print("Start dec time: ", x_values[id_x_start_d])
        i_d_avg = round(y_values[id_x_start_d:id_x_stop_d].mean(),5)
        i_d_max = round(y_values[id_x_start_d:id_x_stop_d].max(),5)
        i_d_min = round(y_values[id_x_start_d:id_x_stop_d].min(),5)
        energy_d = i_d_avg * 3.3 * (x_stop_d-x_start_d)/n_loop
        energy_d = round(energy_d,5)
        # print("Average energy - decryption: ", energy_d, x_start_d, id_x_start_d, x_stop_d, id_x_stop_d, (x_stop_d-x_start_d)/n_loop)

        # # store relevant values
        t_avg_e = round((x_stop_e-x_start_e)/n_loop,5)
        t_avg_d = round((x_stop_d-x_start_d)/n_loop,5)

        # Store results in a text file
        # f = open(output, "w") 
        # f.close()
        # f = open(output, "a")
        # if f.closed:
        #     print("File not open!!")
        # f.write(output)
        # f.write(": \n")
        # f.write("Execution time: \n")
        # f.write("\tEncryption start time:\t\t%f s\n" % x_start_e)
        # f.write("\tEncryption end time:\t\t%f s\n" % x_stop_e)
        # f.write("\tDecryption start time:\t\t%f s\n" % x_start_d)
        # f.write("\tDecryption end time:\t\t%f s\n" % x_stop_d)
        # f.write("\tAverage encryption time:\t%f s\n" % (t_avg_e))
        # f.write("\tAverage decryption time:\t%f s\n\n" % (t_avg_d))
        # f.write("Energy consumption: \n")
        # f.write("\tAverage encryption energy: \t%f J\n" % energy_e)
        # f.write("\tAverage decryption energy: \t%f J\n\n" % energy_d)
        # f.write("Encryption current: \n")
        # f.write("\tMaximum encryption current: %f A\n" % i_e_max)
        # f.write("\tAverage encryption current: %f A\n" % i_e_avg)
        # f.write("\tMinimum encryption current: %f A\n\n" % i_e_min)
        # f.write("Decryption current: \n")
        # f.write("\tMaximum decryption current: %f A\n" % i_d_max)
        # f.write("\tAverage decryption current: %f A\n" % i_d_avg)
        # f.write("\tMinimum decryption current: %f A\n\n" % i_d_min)
        # f.write("Notes: \n")
        # f.write("\tN_LOOP:\t\t\t\t\t\t%d\n" % n_loop)
        # # f.write("\tCalibrated: \t\t\t\t%s\n" % calibrated)
        # f.write("\tNum Calibrations: \t\t\t%d\n" % n_calibrated)
        # f.write("\tCalibration time(s): \t\t")
        # if not calibrated:
        #     f.write("N/A")
        # else:
        #     for t in t_calibrations:
        #         f.write("%.5f s\t" % (round(t/100000,5)))
        #     f.write("\n\tRemoved " + str(n_lcutoff_points) + " samples(s) below " + str(l_cut_off) + "A")
        #     f.write(" and " + str(n_hcutoff_points) + " above " + str(h_cut_off) + "A")
        # f.close
        values = [n_loop, t_avg_e, t_avg_d,  energy_e, energy_d,
          i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated,', '.join(map(str, t_calibrations/100000))]
        # values = [x_start_d - x_stop_e]


    else: 
        # Plotting the graph
        plt.plot(x_values, y_values, linestyle='-')  # You can customize the marker and linestyle as needed
        plt.xlabel('Time (s)')  # Replace 'X Axis Label' with your desired label
        plt.ylabel('Current (A)')  # Replace 'Y Axis Label' with your desired label
        plt.title(data_dir[-15:])  # Replace 'Title of the Graph' with your desired title


        # plt.text(4.28073, 0.047, "Start of program", fontsize=12, ha='center', va='center', color='black')
        plt.grid(True)  # Add gridlines if needed
        plt.show()


def writexl(col, row):
    global values

    # open workbook
    workbook = xl.load_workbook("../Data/Data.xlsx")
    ws = workbook['Power - M4']

    # Write the variable names to the first row
    # for row, var in enumerate(variables, start=row):
    #     ws.cell(row=row, column=col-1, value=var)
    # row = row_start
    for row, var in enumerate(values, start=row):
        ws.cell(row=row, column=col, value=var)
    workbook.save("../Data/Data.xlsx")
if plot_all_bool:
    plot_all()
else:
    if data_obtained:
        t = time.time()
        for opt, data in timestamps.items():
            if "2" in opt:
                run = run3
                opt = opt[:-1]
            else:
                run = run2
            print(opt, run)
            for app, values in data.items():
                if not 'row' in app:
                    # find col
                    col_found = False
                    for col in list(AUT):
                        # print(col.name, app)
                        if app.lower() == col.name.lower():
                            col = col.value
                            col_found = True
                            break
                    if not col_found:
                        if ("romulusnOpt" in app) and ("500x" in app):
                            col = AUT.romulusnOpt_enc_dec_500x.value
                            col_found = True

                    if not col_found:
                        print("Data not stored - col not found!")
                        break
                         
                    print(app)
                    data_dir = "../Data/Power/python_plots/02_csv_dir/" + run + "/" + opt +"/" + app + "/"
                    output = "../Data/Power/python_plots/00_Output/02_run2/" + opt +"/"  + app + ".txt"
                    # print(app)
                    main(data)
                    writexl(col, row)
                    # col +=1
                else:
                    row = values
                    # print("Row ", values)
        print("Took: ", time.time() - t)
    else: 
        main()

